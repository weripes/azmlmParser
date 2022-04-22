import openpyxl

import os

from sync_collector import sync_collector


if __name__ == '__main__':

    print('[:] Скрипт для сбора информации о пользователях azmlm.\n   Данные собираются по перебору ссылок http://azmlm.com/users{число}/')
    # Задавание параметров парсинга с примитивной валидацией
    while True:
        print('\n[?] Какой промежуток страниц пользователей нужно собрать?')
        from_user_id = int(input('[>] От: '))
        to_user_id = int(input('[>] До: '))
        if from_user_id >= 0 and to_user_id > 0 and from_user_id < to_user_id:
            break
        print('\n[!] Недопустимые значения. Повторите попытку')

    if not os.path.isdir('../results'):
        os.mkdir('../results')

    while True:
        filename = input('\n[?] Как назвать файл с результатом?\n[>] ')
        if not os.path.isfile('../results/' + filename + '.xlsx'):
            break
        print('\n[!] Файл с таким названием уже существует. Повторите попытку')

    while True:
        async_or_not = int(input('\n[?] Запустить асинхронно? (0 - Нет, 1 - Да)\n[>] '))
        if async_or_not in (0, 1):
            break
        print('\n[!] Недопустимые значения. Повторите попытку')

    print('\n[*] Сбор данных начялся...')
    # В зависимости от выбора пользователя, импортируется
    # синхронный или асинхронный метод и запускается
    if async_or_not == 0:
        from sync_collector import sync_collector

        users_info = sync_collector(from_user_id=from_user_id, to_user_id=to_user_id)

    else:
        from async_collector import AsyncCollector

        Async_collector = AsyncCollector()
        users_info = Async_collector.start_collector(from_user_id=from_user_id, to_user_id=to_user_id)


    # Создание объекта будущего файла + листа
    wb = openpyxl.Workbook()
    ws = wb.active

    # Заполнение первых верхних ячеек
    ws['A1'] = 'ФИО'
    ws['B1'] = 'Страна, Город'
    ws['C1'] = 'Телефон'
    ws['D1'] = 'Skype'
    ws['E1'] = 'Instagram'
    ws['F1'] = 'Facebook'

    for id, user in enumerate(users_info):
        if not user['FIO']:
            continue
        # "id += 2" - Один на пропуск первых верних ячеек
        # Один потому-что enumerate начинает отсчет с 0
        id += 2 
        ws[f'A{id}'] = user['FIO']
        ws[f'B{id}'] = user['country_and_city']
        ws[f'C{id}'] = user['tel']
        ws[f'D{id}'] = user['skype']
        ws[f'E{id}'] = user['instagram']
        ws[f'F{id}'] = user['facebook']

    wb.save(f'../results/{filename}.xlsx') # Сохранение
    print(f'[*] Данные успешно сохранены. |../results/{filename}.xlsx|')